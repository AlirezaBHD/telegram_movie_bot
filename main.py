import omdb
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pyrogram import Client
import asyncio
import os
import json
from translate import translate_genres, translate_countries

#region Validate Resources

with open("Resource.json", "r") as file:
    config = json.load(file)

telegram_config = config.get("Telegram", {})
omdb_config = config.get("omdb", {})

all_config = {**telegram_config, **omdb_config}

missing_keys = [key for key, value in all_config.items() if not value]

if missing_keys:
    raise ValueError(f"⚠️ The following information is incomplete: {', '.join(missing_keys)}")

#endregion


app = Client("userbot", api_id=config["Telegram"]["api_id"], api_hash=config["Telegram"]["api_hash"])

def create_excel_file(excel_file):
    print("Creating Movie.xlsx")
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Movies"

    headers = ["Name", "Director", "Files & media", "Genre", "Movie or Series", "IMDb", "Rotten Tomatoes",
               "Release Date", "Stars", "Tags", "Downloaded or Not", "Last Watched Episode"]

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    wb.save(excel_file)
    print("Excel File Created.")

class MovieDataHandler:
    def __init__(self, api_key, excel_file):
        omdb.set_default('apikey', api_key)
        if not os.path.exists(excel_file):
            create_excel_file(excel_file)
        self.wb = load_workbook(excel_file)
        self.excel_file = excel_file
        self.sheet = self.wb["Movies"]
        self.no_information = []
        self.done = []

    @staticmethod
    def increase_movie_poster_quality(movie_poster_url):
        if movie_poster_url != "N/A":
            #to use poster image with higher quality
            movie_poster_url = movie_poster_url[:-9] + "FMjpg_UX787_.jpg"
            return movie_poster_url
        return ""

    def get_movie_data(self, movie_name):
        try:
            movie = omdb.get(title=movie_name.strip(), fullplot=True, tomatoes=True)
            if not movie:
                self.no_information.append(movie_name.strip())
                print(f'!! FAILED "{movie_name.strip()}"\n')
            movie["poster"] = self.increase_movie_poster_quality(movie["poster"])
            movie["director"] = "" if movie.get("director") == "N/A" else movie["director"]
            movie["imdb_rating"] = "" if movie.get("imdb_rating") == "N/A" else movie["imdb_rating"]
            movie["runtime"] = movie["runtime"][:-3]
            ratings = movie.get("ratings", [])
            movie["rotten_tomatoes"] = ratings[1]["value"] if len(ratings) > 1 and "value" in ratings[1] else ""
            needed_keys = {"title", "year", "director", "genre", "type", "imdb_rating", "actors", "poster", "runtime",
                           "rated", "country", "writer","rotten_tomatoes"}
            movie = {key: movie[key] for key in needed_keys if key in movie}
            print(movie)

            return movie
        except Exception as e:
            print(f"Error fetching data for '{movie_name}': {e}")
            return None

    async def write_data(self, movie_name):
        movie = self.get_movie_data(movie_name)
        if not movie:
            self.no_information.append(movie_name.strip())
            print(f'!! FAILED "{movie_name.strip()}"\n')
            return
        try:
            confirm = input(f"Did you mean \"{movie["title"]} {movie["year"]}\" by \"{movie.get("director")}\" ? (y/N)")
            if confirm == "y":
                self._update_excel(movie)
                await self._update_telegram(movie)
                self.done.append(f"{movie_name.strip()} - {movie['title']} {movie['year']}" )
                print(f'DONE "{movie["title"]}"\n')
            return
        except Exception as e:
            print(f"Error writing data for '{movie_name}': {e}")

    def _update_excel(self, movie):
        print("---Excel region")

        self.sheet.append([
            movie["title"],
            movie["director"],
            movie["poster"],
            movie.get("genre", ""),
            movie.get("type", ""),
            movie["imdb_rating"],
            movie["rotten_tomatoes"],
            movie.get("year", ""),
            movie.get("actors", "")
        ])
        self.wb.save(self.excel_file)
        print("---Excel endregion")

    @staticmethod
    async def _update_telegram(movie):
        print("---Telegram region")
        information = f'''
{movie["title"]} {movie["year"]}
امتیاز : {movie["imdb_rating"]}

زمان : {movie["runtime"]}دقیقه

ژانر : {translate_genres(movie["genre"])}

رده سنی : {movie["rated"]}

کارگردان : {movie["director"]}

محصول کشور : {translate_countries(movie["country"])}

نویسنده : {movie["writer"]}

بازیگران : {movie["actors"]} ,...
'''
        await app.send_photo(config["Telegram"]["channel_id"], movie["poster"], caption=information)
        print("---Telegram endregion")

    def print_summary(self):
        print(f'Failed to retrieve information for the following movies:\n{self.no_information}\n')
        print(f'Successfully retrieved information for the following movies:\n{self.done}\n')

    def close_excel_book(self):
        self.wb.close()


async def main():
    api_key = config["omdb"]["api_key"]
    excel_file = r'Movie.xlsx'
    handler = MovieDataHandler(api_key, excel_file)

    await app.start()

    while True:
        movie_name = input("Enter the Movie/Series name (or enter '0' to exit):\n")
        if movie_name == "0":
            handler.close_excel_book()
            break
        else:
            await handler.write_data(movie_name)

    handler.print_summary()


if __name__ == "__main__":
    asyncio.run(main())
